import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


df = pd.read_excel('exhibitors.xlsx')

# Clean and split the 'Sectors' column
df['Sectors'] = df['Sectors'].str.replace('\s+', ' ', regex=True).str.strip()
df['Sectors'] = df['Sectors'].str.split(',')


df_exploded = df.explode('Sectors')
# Remove any leading/trailing whitespace from the 'Sectors' entries
df_exploded['Sectors'] = df_exploded['Sectors'].str.strip()

sector_counts = df_exploded['Sectors'].value_counts()

# Group by 'Country' and 'Sectors' and count
country_sector_counts = df_exploded.groupby(['Country', 'Sectors']).size().reset_index(name='Count')

# For each country, get top 3 sectors
top_sectors_per_country = country_sector_counts.sort_values(['Country', 'Count'], ascending=[True, False]).groupby('Country').head(3)


# Plot the top 10 sectors
plt.figure(figsize=(10, 6))
sector_counts.head(10).plot(kind='bar')
plt.title('Top 10 Sectors')
plt.xlabel('Sector')
plt.ylabel('Number of Exhibitors')
plt.tight_layout()
plt.savefig('top_sectors.png')
plt.show()

# Plot the top 10 countries
country_counts = df['Country'].value_counts()
plt.figure(figsize=(10, 6))
country_counts.head(10).plot(kind='bar')
plt.title('Top 10 Countries')
plt.xlabel('Country')
plt.ylabel('Number of Exhibitors')
plt.tight_layout()
plt.savefig('top_countries.png')
plt.show()



def save_graph_to_excel():
    # Load the existing workbook
    wb = load_workbook('exhibitors.xlsx')

    # Create a new sheet for analysis if it doesn't exist
    if 'Analysis' not in wb.sheetnames:
        ws = wb.create_sheet('Analysis')
    else:
        ws = wb['Analysis']

    # Insert the images
    img1 = Image('top_sectors.png')
    img2 = Image('top_countries.png')

    ws.add_image(img1, 'A1')   # Adjust cell positions as needed
    ws.add_image(img2, 'A20')

    # Save the workbook
    wb.save('exhibitors.xlsx')

    # Step 8: Add top sectors per country to Excel
    with pd.ExcelWriter('exhibitors.xlsx', engine='openpyxl', mode='a') as writer:
        top_sectors_per_country.to_excel(writer, sheet_name='Top Sectors per Country', index=False)
